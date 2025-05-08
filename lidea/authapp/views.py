import hashlib
from collections import Counter, defaultdict
from datetime import datetime
import matplotlib.pyplot as plt
from io import BytesIO
import base64
import openpyxl
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from lidea.firebase import db


def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        email_key = email.replace('.', '_')
        hashed_password = hashlib.sha256(password.encode()).hexdigest()

        user_ref = db.reference(f'users/{email_key}')
        user_data = user_ref.get()

        if not user_data:
            return render(request, 'login.html', {'error': 'Пользователь не найден.'})

        if user_data['user_password'] != hashed_password:
            return render(request, 'login.html', {'error': 'Неверный пароль.'})

        if user_data.get('role_name') != 'бухгалтер':
            return render(request, 'error.html', {'message': 'Доступ разрешён только бухгалтерам.'})

        # Всё успешно: сохраняем в сессию
        request.session['user_name'] = user_data.get('user_name')
        request.session['user_email'] = user_data.get('user_email')
        request.session['role_name'] = user_data.get('role_name')

        return redirect('success')

    return render(request, 'login.html')

def success_view(request):
    if request.session.get('role_name') != 'бухгалтер':
        return redirect('login')

    user_name = request.session.get('user_name')

    data_ref = db.reference('production')
    raw_data = data_ref.get() or []

    def parse_date(date_str):
        return datetime.strptime(date_str, '%d.%m.%Y')

    aggregated_data = defaultdict(lambda: {'quantity': 0, 'product_names': set()})

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    for item in raw_data:
        try:
            item_date_str = item.get('date_production')
            item_date = parse_date(item_date_str)

            if start_date_str and end_date_str:
                start = parse_date(start_date_str)
                end = parse_date(end_date_str)
                if not (start <= item_date <= end):
                    continue

            quantity = int(item.get('product_quantity', 0))
            product_name = item.get('product_name', 'Неизвестно')
            aggregated_data[item_date]['quantity'] += quantity
            aggregated_data[item_date]['product_names'].add(product_name)

        except Exception as e:
            print("Ошибка обработки элемента:", item, e)
            continue

    chart_data = [
        {
            'date': date.strftime('%d.%m.%Y'),
            'quantity': data['quantity'],
            'product_names': ', '.join(data['product_names'])
        }
        for date, data in sorted(aggregated_data.items())
    ]

    return render(request, 'success.html', {
        'user_name': user_name,
        'data': chart_data,
        'start_date': start_date_str,
        'end_date': end_date_str
    })

def logout_view(request):
    # Очищаем сессию
    request.session.flush()
    return redirect('login')

def rating_view(request):
    # Получаем все одобренные отзывы
    reviews_data = db.reference('product_review').get() or {}
    products_data = db.reference('products').get() or []

    product_info_by_name = {
        product.get('product_name'): {
            'photo': product.get('product_photo'),
            'description': product.get('product_description')
        }
        for product in products_data
    }

    # Формируем список отзывов с фото и описанием
    reviews = []
    for review in reviews_data.values():
        if not review.get('approved'):
            continue
        name = review.get('product_name')
        reviews.append({
            'product_name': name,
            'product_photo': product_info_by_name.get(name, {}).get('photo', ''),
            'product_description': product_info_by_name.get(name, {}).get('description', ''),
            'rating': review.get('product_evaluation', 0)
        })

    # Сортировка по параметру (если параметр есть)
    sort = request.GET.get('sort')
    if sort == 'high':
        reviews.sort(key=lambda x: x['rating'], reverse=True)
    elif sort == 'low':
        reviews.sort(key=lambda x: x['rating'])
    elif sort == 'none':
        reviews = [r for r in reviews if not r['rating']]

    return render(request, 'rating.html', {'reviews': reviews, 'sort': sort})

def accountant_orders_view(request):
    if request.session.get('role_name') != 'бухгалтер':
        return redirect('login')

    orders_data = db.reference('orders').get() or {}
    products_data = db.reference('products_in_order').get() or {}

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, '%d.%m.%Y')
        except:
            return None

    start = parse_date(start_date_str) if start_date_str else None
    end = parse_date(end_date_str) if end_date_str else None

    # Группируем продукты по заказам
    products_by_order = defaultdict(list)
    for p in products_data.values():
        products_by_order[p['id_orders']].append(p)

    orders = []
    order_count = 0
    total_cost = 0
    total_delivery = 0
    product_counter = Counter()
    status = request.GET.get('status')

    for order_id, order in orders_data.items():
        order_date_str = order.get('order_date')
        order_date = parse_date(order_date_str)
        if not order_date:
            continue

        if (start and order_date < start) or (end and order_date > end):
            continue

        if status and order.get('name_order_status') != status:
            continue

        order_products = products_by_order.get(order_id, [])
        for prod in order_products:
            product_counter[prod['product_name']] += int(prod.get('quantity_products_order', 0))

        orders.append({
            'id': order_id,
            'order_date': order_date_str,
            'order_address': order.get('order_address'),
            'id_courier': order.get('id_courier'),
            'name_order_status': order.get('name_order_status'),
            'order_cost': order.get('order_cost'),
            'order_delivery_cost': order.get('order_delivery_cost'),
            'products': order_products
        })

        order_count += 1
        total_cost += int(order.get('order_cost', 0))
        total_delivery += int(order.get('order_delivery_cost', 0))

    # Список популярных товаров
    popular_products = [{'product_name': name, 'total_quantity': quantity}
                        for name, quantity in product_counter.most_common(10)]

    return render(request, 'accounting.html', {
        'orders': orders,
        'order_count': order_count,
        'status': status,
        'total_cost': total_cost,
        'total_delivery': total_delivery,
        'total_all': total_cost + total_delivery,
        'popular_products': popular_products,
        'start_date': start_date_str or '—',
        'end_date': end_date_str or '—'
    })



def export_orders_to_excel(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    courier = request.GET.get('courier')
    status = request.GET.get('status')

    orders_data = db.reference('orders').get() or {}
    products_data = db.reference('products_in_order').get() or {}

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = ['ID заказа', 'Дата', 'Адрес', 'Курьер', 'Статус', 'Товары', 'Цена', 'Доставка']
    ws.append(headers)

    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, '%d.%m.%Y')
        except Exception:
            return None

    all_orders = []
    for order_id, order in orders_data.items():
        order_date = parse_date(order.get('order_date'))
        if not order_date:
            continue

        if start_date_str:
            start_date = parse_date(start_date_str)
            if start_date and order_date < start_date:
                continue
        if end_date_str:
            end_date = parse_date(end_date_str)
            if end_date and order_date > end_date:
                continue

        if courier and order.get('id_courier') != courier:
            continue
        if status and order.get('name_order_status') != status:
            continue

        all_orders.append((order_id, order, order_date))

    # Сортировка по дате
    all_orders.sort(key=lambda x: x[2])

    for order_id, order, order_date in all_orders:
        order_address = order.get('order_address', 'Неизвестно')
        courier_id = order.get('id_courier')
        courier_id = courier_id if courier_id else 'Не назначен'
        order_status = order.get('name_order_status', 'Неизвестен')
        order_cost = order.get('order_cost', 0)
        delivery_cost = order.get('order_delivery_cost', 0)

        order_products = [
            {
                'product_name': product.get('product_name', 'Без имени'),
                'product_price': product.get('product_price', 0),
                'quantity': product.get('quantity_products_order', 0)
            }
            for product in products_data.values()
            if isinstance(product, dict) and product.get('id_orders') == order_id
        ]

        if order_products:
            products_list = ', '.join(
                f"{prod['product_name']} ({prod['quantity']} шт., {prod['product_price']} руб.)"
                for prod in order_products
            )
        else:
            products_list = 'Нет товаров'

        ws.append([
            order_id,
            order.get('order_date', 'Неизвестно'),
            order_address,
            courier_id,
            order_status,
            products_list,
            order_cost,
            delivery_cost
        ])

    # Применяем стили к заголовкам
    for col_num, col_title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Автоширина столбцов
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(15, max_length + 2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
    wb.save(response)
    return response



import json  # Вверху файла

def supply_history_view(request):
    if request.session.get('role_name') != 'бухгалтер':
        return redirect('login')

    supply_data = db.reference('supply').get() or {}

    supply_name_filter = request.GET.get('supply_name')
    product_name_filter = request.GET.get('product_name')
    min_quantity_filter = request.GET.get('min_quantity')

    aggregated_data = defaultdict(lambda: {'total_quantity': 0, 'products': defaultdict(int)})
    chart_dict = defaultdict(int)  # для объединения товаров по ключу (поставщик, товар)

    for item in supply_data:
        try:
            supply_name = item.get('supply_name', 'Неизвестно')
            product_name = item.get('product_name', 'Неизвестно')
            quantity = int(item.get('product_quantity', 0))

            if supply_name_filter and supply_name != supply_name_filter:
                continue
            if product_name_filter and product_name != product_name_filter:
                continue
            if min_quantity_filter and quantity < int(min_quantity_filter):
                continue

            # Группируем данные для таблицы
            aggregated_data[supply_name]['total_quantity'] += quantity
            aggregated_data[supply_name]['products'][product_name] += quantity

            # Объединяем данные по (поставщик, товар)
            key = f"{supply_name} — {product_name}"
            chart_dict[key] += quantity

        except Exception as e:
            print("Ошибка обработки поставки:", item, e)
            continue

    # Формируем список данных для диаграммы
    chart_items = [{'label': k, 'quantity': v} for k, v in chart_dict.items()]

    chart_data = [
        {
            'supply_name': supply_name,
            'total_quantity': data['total_quantity'],
            'products': [
                {'product_name': product, 'quantity': quantity}
                for product, quantity in data['products'].items()
            ]
        }
        for supply_name, data in aggregated_data.items()
    ]

    return render(request, 'supply_history.html', {
        'chart_data': chart_data,
        'chart_items_json': json.dumps(chart_items, ensure_ascii=False),
        'supply_name_filter': supply_name_filter or '',
        'product_name_filter': product_name_filter or '',
        'min_quantity_filter': min_quantity_filter or ''
    })



def generate_excel_report(request):
    if request.session.get('role_name') != 'бухгалтер':
        return redirect('login')

    # Получаем данные о поставках из Firebase
    supply_data = db.reference('supply').get() or {}

    # Получаем параметры для фильтрации из GET-запроса
    supply_name_filter = request.GET.get('supply_name')  # Фильтр по имени поставщика
    product_name_filter = request.GET.get('product_name')  # Фильтр по имени товара
    min_quantity_filter = request.GET.get('min_quantity')  # Минимальное количество товара

    # Группируем данные по поставщикам и товарам
    aggregated_data = defaultdict(lambda: {'total_quantity': 0, 'products': defaultdict(int)})

    for item in supply_data:
        try:
            # Извлекаем данные о поставке
            supply_name = item.get('supply_name', 'Неизвестно')
            product_name = item.get('product_name', 'Неизвестно')
            quantity = int(item.get('product_quantity', 0))

            # Фильтруем по имени поставщика, если задано
            if supply_name_filter and supply_name != supply_name_filter:
                continue

            # Фильтруем по имени товара, если задано
            if product_name_filter and product_name != product_name_filter:
                continue

            # Фильтруем по минимальному количеству товара, если задано
            if min_quantity_filter and quantity < int(min_quantity_filter):
                continue

            # Группируем данные по поставщикам и товарам
            aggregated_data[supply_name]['total_quantity'] += quantity
            aggregated_data[supply_name]['products'][product_name] += quantity

        except Exception as e:
            print("Ошибка обработки поставки:", item, e)
            continue

    # Создание Excel-файла
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по поставкам"

    # Запись заголовков
    ws.append(['Поставщик', 'Товар', 'Количество товара', 'Общее количество'])

    # Заполнение Excel данными
    for supply_name, data in aggregated_data.items():
        for product_name, quantity in data['products'].items():
            ws.append([supply_name, product_name, quantity, data['total_quantity']])

    # Создание ответа с Excel файлом
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="supply_report.xlsx"'

    # Сохранение файла в ответ
    wb.save(response)
    return response


